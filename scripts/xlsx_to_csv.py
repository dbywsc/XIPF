"""Convert CCPC/ICPC XLSX standings to standard CSV format.

Usage: python scripts/xlsx_to_csv.py contest.xlsx

Outputs tab-separated CSV(s) with computed school rankings.
For multi-division events (邀请赛 + 省赛), generates separate files.
"""

import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)


CITY_NAMES = [
    "北京", "上海", "广州", "深圳", "杭州", "南京", "武汉", "成都", "重庆",
    "西安", "长沙", "郑州", "济南", "青岛", "合肥", "南昌", "福州", "厦门",
    "沈阳", "大连", "哈尔滨", "长春", "昆明", "南宁", "贵阳", "兰州",
    "天津", "石家庄", "太原", "苏州", "徐州", "扬州", "宁波", "温州",
    "秦皇岛", "三亚", "桂林", "珠海", "佛山", "东莞", "中山", "惠州",
    "汕头", "绍兴", "芜湖", "洛阳", "开封", "湘潭", "株洲", "衡阳",
    "绵阳", "咸阳", "自贡", "包头", "齐齐哈尔",
]

PINYIN_CITIES = {
    "qinhuangdao": "秦皇岛", "beijing": "北京", "shanghai": "上海",
    "guangzhou": "广州", "shenzhen": "深圳", "hangzhou": "杭州",
    "nanjing": "南京", "wuhan": "武汉", "chengdu": "成都", "chongqing": "重庆",
    "xian": "西安", "changsha": "长沙", "zhengzhou": "郑州", "jinan": "济南",
    "qingdao": "青岛", "hefei": "合肥", "nanchang": "南昌", "fuzhou": "福州",
    "xiamen": "厦门", "shenyang": "沈阳", "dalian": "大连", "haerbin": "哈尔滨",
    "changchun": "长春", "kunming": "昆明", "nanning": "南宁", "guiyang": "贵阳",
    "lanzhou": "兰州", "tianjin": "天津", "shijiazhuang": "石家庄", "taiyuan": "太原",
    "suzhou": "苏州", "xuzhou": "徐州", "yangzhou": "扬州", "ningbo": "宁波",
    "wenzhou": "温州", "dongguan": "东莞",
}


def load_org_name_map() -> dict:
    """Load English→Chinese organization name mapping."""
    import json
    map_path = Path(__file__).resolve().parent.parent / "org_names_map.json"
    if map_path.exists():
        with open(map_path, encoding="utf-8") as f:
            return json.load(f)
    return {}

ORG_NAME_MAP = None

def translate_org(name: str) -> str:
    global ORG_NAME_MAP
    if ORG_NAME_MAP is None:
        ORG_NAME_MAP = load_org_name_map()
    return ORG_NAME_MAP.get(name.strip(), name)


def detect_city(filename: str) -> str:
    lower = filename.lower()
    for c in CITY_NAMES:
        if c in lower:
            return c
    for pinyin, chinese in PINYIN_CITIES.items():
        if pinyin in lower:
            return chinese
    return "unknown"


def detect_division(filename: str) -> str:
    lower = filename.lower()
    if "invitational" in lower or "邀请赛" in lower:
        return "invitational"
    if "provincial" in lower or "省赛" in lower:
        return "provincial"
    return ""


def parse_medal(rank_str: str) -> str:
    if not rank_str:
        return ""
    s = str(rank_str).strip().lower()
    if "gold" in s or "金奖" in s or "金牌" in s:
        return "金奖"
    if "silver" in s or "银奖" in s or "银牌" in s:
        return "银奖"
    if "bronze" in s or "铜奖" in s or "铜牌" in s:
        return "铜奖"
    return ""


def parse_members(value: str) -> list:
    if not value:
        return ["", "", ""]
    names = []
    for part in str(value).split(","):
        part = part.strip()
        if not part:
            continue
        if "教练" in part:
            continue
        names.append(part)
    while len(names) < 3:
        names.append("")
    return names[:3]


def read_main_sheet(ws, rank_col: int = 0) -> list:
    """Read a sheet and return list of team dicts.
    rank_col: 0 for invitational (col A), 1 for provincial (col B)."""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        return []

    header_row = -1
    for i, row in enumerate(rows):
        if not row:
            continue
        first_cells = [str(c).strip().lower() if c else "" for c in row[:8]]
        if any(kw in " ".join(first_cells) for kw in ["rank", "organization", "name", "team", "school", "成员"]):
            header_row = i
            break
    if header_row < 0:
        header_row = 1

    headers = [str(c).strip() if c else "" for c in rows[header_row]]
    data_start = header_row + 1

    # Auto-detect column positions from headers
    def find_col(keywords: list) -> int:
        for i, h in enumerate(headers):
            hl = h.lower()
            if any(kw in hl for kw in keywords):
                return i
        return -1

    col_rank_val = rank_col  # 0 for invitational (col A), 1 for provincial (col B)
    col_org_default = find_col(["organization", "学校", "org"])
    col_team = find_col(["name", "team", "队伍", "队名"])
    col_org = col_org_default if col_org_default >= 0 else (find_col(["official"]) + 1 if find_col(["official"]) >= 0 else 5)  # org is usually right after official
    member_col = find_col(["member", "队员", "成员"])
    col_official = find_col(["official"])

    teams = []
    for row in rows[data_start:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        col_a = str(row[col_rank_val]).strip() if len(row) > col_rank_val and row[col_rank_val] else ""
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        col_d = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        organization = translate_org(str(row[col_org]).strip()) if col_org >= 0 and len(row) > col_org else ""
        team_name = str(row[col_team]).strip() if col_team >= 0 and len(row) > col_team else ""
        official_marker = str(row[col_official]).strip() if col_official >= 0 and len(row) > col_official else ""
        raw_members = str(row[member_col]).strip() if member_col >= 0 and len(row) > member_col else ""
        members = parse_members(raw_members)
        unofficial = col_a == "*" or official_marker == "*"

        teams.append({
            "rank_str": col_a,
            "overall_rank": int(col_b) if col_b.isdigit() else 0,
            "organization": organization,
            "team_name": team_name,
            "members": members,
            "unofficial": unofficial,
            "girl": "女队" in col_d,
        })
    return teams


def compute_school_ranks(teams: list) -> dict:
    org_best = {}
    for t in teams:
        if t["unofficial"]:
            continue
        org = t["organization"]
        if not org:
            continue
        if org not in org_best or t["rank_num"] < org_best[org]:
            org_best[org] = t["rank_num"]
    sorted_orgs = sorted(org_best.items(), key=lambda x: x[1])
    return {org: i + 1 for i, (org, _) in enumerate(sorted_orgs)}


def write_csv(teams: list, org_ranks: dict, out_path: Path):
    """Write a single CSV file from processed teams."""
    output_rows = []
    output_rows.append([
        "Rank", "Organization Rank", "Organization", "Team",
        "Member1", "Member2", "Member3",
        "Unofficial", "Girl", "Prize",
    ])
    for t in teams:
        medal = parse_medal(t["rank_str"]) if not t["unofficial"] else ""
        rank_display = str(t["rank_num"]) if not t["unofficial"] else "*"
        org_rank = org_ranks.get(t["organization"], "")
        output_rows.append([
            rank_display, str(org_rank), t["organization"], t["team_name"],
            t["members"][0], t["members"][1], t["members"][2],
            "Y" if t["unofficial"] else "N", "Y" if t["girl"] else "N", medal,
        ])
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerows(output_rows)

    off = [t for t in teams if not t["unofficial"]]
    unoff = [t for t in teams if t["unofficial"]]
    print(f"  -> {out_path.name}: {len(off)} official + {len(unoff)} unofficial, {len(org_ranks)} orgs")


def convert(filepath: Path, output: Path = None):
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    sheet_names = wb.sheetnames
    filename = filepath.stem
    city = detect_city(filename)
    division = detect_division(filename)

    # Detect division sheets (e.g., "邀请赛", "河北省赛")
    division_sheets = []
    for name in sheet_names:
        lower = name.strip()
        if any(kw in lower for kw in ["邀请赛", "invitational"]):
            division_sheets.append((name, "invitational"))
        elif any(kw in lower for kw in ["省赛", "provincial"]):
            division_sheets.append((name, "provincial"))

    if division_sheets:
        # If only one division type found (e.g., only provincial), also process Main as the other type
        div_types = {d[1] for d in division_sheets}
        if "invitational" not in div_types:
            division_sheets.append(("Main", "invitational"))
        elif "provincial" not in div_types:
            division_sheets.append(("Main", "provincial"))
        print(f"City: {city}, Divisions: {[d[1] for d in division_sheets]}")
        for sheet_name, div_suffix in division_sheets:
            rank_col = 0 if div_suffix == "invitational" else 1
            teams = read_main_sheet(wb[sheet_name], rank_col)
            if not teams:
                continue
            official_count = 0
            for t in teams:
                if t["unofficial"]:
                    t["rank_num"] = 0
                else:
                    official_count += 1
                    t["rank_num"] = official_count
            org_ranks = compute_school_ranks(teams)
            out_path = output or filepath.parent / f"{city}_{div_suffix}.csv"
            write_csv(teams, org_ranks, out_path)
    else:
        # Single contest
        ws = wb["Main"] if "Main" in sheet_names else wb[sheet_names[0]]
        teams = read_main_sheet(ws)
        if teams:
            official_count = 0
            for t in teams:
                if t["unofficial"]:
                    t["rank_num"] = 0
                else:
                    official_count += 1
                    t["rank_num"] = official_count
            org_ranks = compute_school_ranks(teams)
            slug = city
            if division:
                slug += f"_{division}"
            out_path = output or filepath.parent / f"{slug}.csv"
            write_csv(teams, org_ranks, out_path)

    wb.close()


def main():
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} contest.xlsx [output.csv]")
        print("\nConverts CCPC/ICPC XLSX standings to tab-separated CSV.")
        sys.exit(1)
    filepath = Path(sys.argv[1]).resolve()
    if not filepath.exists():
        print(f"File not found: {filepath}")
        sys.exit(1)
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    convert(filepath, output)


if __name__ == "__main__":
    main()
